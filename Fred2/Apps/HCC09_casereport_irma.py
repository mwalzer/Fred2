import os
from collections import defaultdict
from Fred2.IO.ADBAdapter import EAdapterFields, EIdentifierTypes
import Fred2.Core.Generator as g
from Fred2.Core.Allele import Allele
from Fred2.Core.Peptide import Protein
from Fred2.Core.Result import EpitopePredictionResult
from Fred2.EpitopePrediction import EpitopePredictorFactory
from Fred2.IO.MartsAdapter import MartsAdapter
from Fred2.IO.RefSeqAdapter import RefSeqAdapter
from Fred2.IO.ADBAdapter import EAdapterFields
import re
import csv
import math
import pandas as pd
import xlsxwriter
import itertools
from openpyxl import load_workbook
from pandas import ExcelWriter

from Fred2.Core import Variant
from Fred2.Core import Transcript
from Fred2.Core import VariationType
from Fred2.Core import MutationSyntax

import logging


def create_transcript_column_value(pep):
    #create from first index which is a tuple in Fred2 prediction result dataframes
    return ','.join([x.transcript_id for x in pep[0].get_all_transcripts()])


def create_mutationsyntax_column_value(pep):
    #create from first index which is a tuple in Fred2 prediction result dataframes
    v = [x.vars.values() for x in pep[0].get_all_transcripts()]
    vf = list(itertools.chain.from_iterable(v))
    c = [x.coding.values() for x in vf]
    cf = list(itertools.chain.from_iterable(c))
    return ','.join([y.aaMutationSyntax for y in cf])


def create_variationfilelinenumber_column_value(pep):
    #create from first index which is a tuple in Fred2 prediction result dataframes
    v = [x.vars.values() for x in pep[0].get_all_transcripts()]
    vf = list(itertools.chain.from_iterable(v))
    return ','.join([y.id+1 for y in vf])


def create_gene_column_value(pep):
    #create from first index which is a tuple in Fred2 prediction result dataframes
    v = [x.vars.values() for x in pep[0].get_all_transcripts()]
    vf = list(itertools.chain.from_iterable(v))
    return ','.join([y.gene for y in vf])


def convert_logscore_to_nM(val):
    return math.pow(50000, 1.0-val)


def add_xslx_sheet(xlsx_file, sheetname, dataframe):
    book = load_workbook(xlsx_file)
    writer = pd.ExcelWriter(xlsx_file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    dataframe.to_excel(writer, sheetname)
    writer.save()

pd.set_option('precision',4)

logging.basicConfig(format='%(levelname)s:%(message)s', level=logging.INFO)
logging.info('Starting HCC009 casereport variant integration and prediction')

target_alleles_set = set()
#A*03, A*29, B*07, B*35, C*04, C*15
with open("/home/walzer/bwSyncAndShare/hcc009casereport/hcc009.allele", 'r') as handle:
    for line in handle:
        target_alleles_set.add(Allele(line.strip().upper()))

type_mapper = {'missense_variant': VariationType.SNP,
                'stop_gained&splice_region_variant': VariationType.SNP,
                   ('nonframeshift', 'deletion'): VariationType.DEL,
                   ('frameshift', 'deletion'): VariationType.FSDEL,
                   ('nonframeshift', 'insertion'): VariationType.INS,
                   ('frameshift', 'insertion'): VariationType.FSINS}

# ttn = EpitopePredictorFactory('netmhc', version="3.4")
ttn = EpitopePredictorFactory('netMHC', version="3.0a")
# tts = EpitopePredictorFactory('syfpeithi', version="1.0.2015")
tts = EpitopePredictorFactory('syfpeithi', version="original")
# ttp = EpitopePredictorFactory('netmhcpan', version="2.8")
ttp = EpitopePredictorFactory('netmhcpan', version="2.4")

ma = MartsAdapter(biomart="http://ensembl.org")
#ma = MartsAdapter(biomart="http://grch37.ensembl.org")
refseq = RefSeqAdapter(mrna_file="/home/walzer/dbs/RefSeq_human.rna.fna")

###
files = ["/home/walzer/bwSyncAndShare/hcc009casereport/GS100038-GS120152_22.01.2016.csv",
         "/home/walzer/bwSyncAndShare/hcc009casereport/GS120154-GS120152_22.01.2015.csv",
         "/home/walzer/bwSyncAndShare/hcc009casereport/GS130348_01-GS120152_22.01.2015.csv"]

for file in files:
    logging.info("Processing file %s"%(file))
    vars = list()
    lines = list()

    #RE = re.compile("(\w+):([\w.]+):(\w+):\w*:exon(\d+)\D*\d*:(c.\D*(\d+)\D*):(p.\D*(\d+)\D*),")
    RE = re.compile("(\w+):([\w.]+):(\w+):\w*:exon(\d+)\D*\d*:(c.\D*(\d+)\D*):(p.\D*(\d+)\w*)")

    with open(file, "r") as f:
    #with open("/home/walzer/bwSyncAndShare/hcc009casereport/GS100038-GS120152_22.01.2016.csv", "r") as f:
    #with open("/home/walzer/bwSyncAndShare/hcc009casereport/GS120154-GS120152_22.01.2015.csv", "r") as f:
    #with open("/home/walzer/bwSyncAndShare/hcc009casereport/GS130348_01-GS120152_22.01.2015.csv", "r") as f:
        reader = csv.DictReader(f, delimiter='\t', quoting=csv.QUOTE_NONE)
        #print reader.fieldnames
        lines = [x for x in reader]

    lineswithvariants = set()
    for mut_id, line in enumerate(lines):
        genome_start = line["start"]
        genome_stop = line["end"]
        chrom = line["#chr"]
        ref = line["ref"]
        alt = line["obs"]
        zygos = "hom"
        annots = RE.findall(line["coding_and_splicing"])
        for annot in annots:
            gene, nm_id, mut_type, exon, trans_coding, trans_pos, prot_coding, prot_start = annot
            if "stop_gained" not in mut_type:
                nm_id = nm_id.split(".")[0]
                #print nm_id
                coding = dict()
                coding[nm_id] = MutationSyntax(nm_id, int(trans_pos)-1, int(prot_start)-1, trans_coding, prot_coding)
                vars.append(
                        Variant(mut_id, type_mapper.get(mut_type, VariationType.UNKNOWN), chrom, int(genome_start),
                                ref.upper(), alt.upper(), coding, zygos == "hom", isSynonymous=False))
                vars[-1].gene = gene
                lineswithvariants.add(mut_id)

    for linenr,line in enumerate(lines):
        if linenr not in lineswithvariants:
            logging.warning("No Variant from line: "+'\t'.join(line.values()))

    ts = list()
    #using a tweaked generator that takes another sequence source if the sequence is too short in respect to the given variants
    #in this case a newer/older sequence from mart in respect to what was given as reference in the annotation process
    t = g.generate_transcripts_from_variants(vars, ma, id_type=EIdentifierTypes.REFSEQ, sequenceInject=refseq)
    ts = [x for x in t]
    p = g.generate_proteins_from_transcripts(ts, to_stop=True)
    ps = [x for x in p]

    workbook = xlsxwriter.Workbook(file+".predictions.xlsx")
    workbook.close()
    workbook = xlsxwriter.Workbook(file+".predictions.filtered.xlsx")
    workbook.close()
    wpred = ExcelWriter(file+".predictions.xlsx")
    wfilt = ExcelWriter(file+".predictions.filtered.xlsx")
    for peplen in range(8,13):
        logging.info("Prediction of length %i"%(peplen))
        e = g.generate_peptides_from_protein(ps, peplen, only_variants=True)
        es = [x for x in e]
        #if not os.path.isfile(file+".predictions.xlsx"):
        try:
            preds_n = ttn.predict(es, alleles=list(target_alleles_set))
            preds_n_f = pd.DataFrame()
            if not preds_n.empty:
                preds_n = preds_n.applymap(convert_logscore_to_nM)
                preds_n_f = preds_n.loc[(preds_n<500).any(axis=1)]

                preds_n['Transcript'] = preds_n.index.map(create_transcript_column_value)
                preds_n.rename(columns=lambda x: str(x), inplace=True)
                preds_n.index = preds_n.index.droplevel(1)
                preds_n.rename(index=lambda x: str(x), inplace=True)

                if not preds_n_f.empty:
                    preds_n_f['Transcript'] = preds_n_f.index.map(create_transcript_column_value)
                    preds_n_f.rename(columns=lambda x: str(x), inplace=True)
                    preds_n_f.index = preds_n_f.index.droplevel(1)
                    preds_n_f.rename(index=lambda x: str(x), inplace=True)

            logging.info("unfiltered predicted peptides length %i"%(len(preds_n)))
            logging.info("filtered predicted peptides length %i"%(len(preds_n_f)))


            #preds_n.to_csv(file+".netmhc"+str(peplen)+".csv", sep="\t")
            preds_n.to_excel(wpred, str(ttn.name)+str(ttn.version)+"_"+str(peplen),float_format="%.4f")
            preds_n_f.to_excel(wfilt, str(ttn.name)+str(ttn.version)+"_"+str(peplen),float_format="%.4f")

        except Exception as e:
            logging.error("something went wrong with the netMHC prediction on file %s:"%file)
            logging.error(str(e),e)
        try:
            if peplen<12:
                preds_p = ttp.predict(es, alleles=list(target_alleles_set))
                preds_p_f = pd.DataFrame()
                if not preds_p.empty:
                    preds_p = preds_p.applymap(convert_logscore_to_nM)
                    preds_p_f = preds_p.loc[(preds_p<500).any(axis=1)]

                    preds_p['Transcript'] = preds_p.index.map(create_transcript_column_value)
                    preds_p.rename(columns=lambda x: str(x), inplace=True)
                    preds_p.index = preds_p.index.droplevel(1)
                    preds_p.rename(index=lambda x: str(x), inplace=True)

                    if not preds_p_f.empty:
                        preds_p_f['Transcript'] = preds_p_f.index.map(create_transcript_column_value)
                        preds_p_f.rename(columns=lambda x: str(x), inplace=True)
                        preds_p_f.index = preds_p_f.index.droplevel(1)
                        preds_p_f.rename(index=lambda x: str(x), inplace=True)

                #preds_p.to_csv(file+".netpan"+str(peplen)+".csv", sep="\t")
                preds_p.to_excel(wpred, str(ttp.name)+str(ttp.version)+"_"+str(peplen),float_format="%.4f")
                preds_p_f.to_excel(wfilt, str(ttp.name)+str(ttp.version)+"_"+str(peplen),float_format="%.4f")
            else:
                logging.info("No NetMHCpan prediction for size >11")

        except Exception as e:
            logging.error("something went wrong with the netMHCpan prediction on file %s:"%file)
            logging.error(str(e),e)
        try:
            preds_s = tts.predict(es, alleles=list(target_alleles_set))
            preds_s_f = pd.DataFrame()
            if not preds_s.empty:
                preds_s = preds_s * 100
                preds_s_f = preds_s.loc[(preds_s>50).any(axis=1)]

                preds_s['Transcript'] = preds_s.index.map(create_transcript_column_value)
                preds_s.rename(columns=lambda x: str(x), inplace=True)
                preds_s.index = preds_s.index.droplevel(1)
                preds_s.rename(index=lambda x: str(x), inplace=True)

                if not preds_s_f.empty:
                    preds_s_f['Transcript'] = preds_s_f.index.map(create_transcript_column_value)
                    preds_s_f.rename(columns=lambda x: str(x), inplace=True)
                    preds_s_f.index = preds_s_f.index.droplevel(1)
                    preds_s_f.rename(index=lambda x: str(x), inplace=True)

            #preds_s.to_csv(file+".syfpeithi"+str(peplen)+".csv", sep="\t")
            preds_s.to_excel(wpred, str(tts.name)+str(tts.version)+"_"+str(peplen),float_format="%.4f")
            preds_s_f.to_excel(wfilt, str(tts.name)+str(tts.version)+"_"+str(peplen),float_format="%.4f")

        except Exception as e:
            logging.error("something went wrong with the syfpeithi prediction on file %s:"%file)
            logging.error(str(e),e)

        # df = EpitopePredictionResult(preds_p_f).merge_results([EpitopePredictionResult(x) for x in [preds_n_f, preds_s_f]])
        # df.to_excel(wfilt, "combined"+"_"+str(peplen),float_format="%.4f")


    wpred.save()
    wfilt.save()